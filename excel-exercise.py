import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

# 載入活頁簿
wb = openpyxl.load_workbook('example.xlsx')
print(wb.get_sheet_names())
# 取得工作頁
sheet = wb.get_sheet_by_name('Sheet1')
# 取得儲存格
cell = sheet['B1']
# row 列數 column 欄數 value 內容值
print('Row ' + str(cell.row) + ', Column ' + str(cell.column) + ' is ' + cell.value)
print('Cell ' + cell.coordinate + ' is ' + cell.value)
print("======我是切割線======")
# 利用cell方法指定儲存格位置
for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)
print("======我是切割線======")
# 確定工作表的內容大小
print(sheet.max_row, '|', sheet.max_column)
print("======我是切割線======")
# 欄的字母和數字之間的轉換
print(get_column_letter(900))
print(column_index_from_string('AA'))
print("======我是切割線======")
# 指定A1到C3的矩形範圍, 並進行迴圈以存取所有的儲存格
for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
print("======我是切割線======")
# 指定欄或列的儲存格中的值
for cellObj in list(sheet.columns)[1]:
    print(cellObj.value)