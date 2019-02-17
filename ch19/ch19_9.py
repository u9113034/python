# ch19_9.py
import openpyxl

fn = 'sales.xlsx'
wb = openpyxl.load_workbook(fn, data_only=True)
#ws = wb.get_sheet_by_name('2020Q1')                 # 目前工作表2020Q1
ws = wb['2020Q1']
for i in range(1, ws.max_column+1):                 # column做索引增值
    print(ws.cell(column=i, row=5).value, end=' ')  # row=5, 索引不變
  




