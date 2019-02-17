# ch19_19.py
import openpyxl

wb = openpyxl.Workbook()                            # 建立空白的活頁簿
print("所有工作表名稱 = ", wb.sheetnames)    # 列印所有工作表
wb.create_sheet()                                   # 建立新工作表
print("所有工作表名稱 = ", wb.sheetnames)    # 列印所有工作表
ws = wb.active                                # 取得目前工作表
print("目前工作表名稱 = ", ws.title)                # 列印目前工作表
wb.save('out19_19.xlsx')                            # 將活頁簿儲存





