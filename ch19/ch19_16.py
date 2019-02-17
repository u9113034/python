# ch19_16.py
import openpyxl
#from openpyxl.utils import get_column_letter, column_index_from_string

fn = 'sales.xlsx'
wb = openpyxl.load_workbook(fn, data_only=True)
#ws = wb.get_sheet_by_name('2020Q1')             # 目前工作表2020Q1
ws = wb['2020Q1']
for row in ws['A3':'E6']:
    for cell in row:
        print(cell.value, end=' ')
    print()




