# ch19_20.py
import openpyxl

wb = openpyxl.Workbook()                            # 建立空白的活頁簿
print("所有工作表名稱 = ", wb.sheetnames)    # 列印所有工作表
wb.create_sheet()                                   # 建立新工作表
print("所有工作表名稱 = ", wb.sheetnames)    # 列印所有工作表
wb.create_sheet(index=0, title='First sheet')       # 第一個工作表
print("所有工作表名稱 = ", wb.sheetnames)    # 列印所有工作表
wb.create_sheet(index=2, title='Third sheet')       # 第三個工作表
print("所有工作表名稱 = ", wb.sheetnames)    # 列印所有工作表
wb.save('out19_20.xlsx')                            # 將活頁簿儲存





