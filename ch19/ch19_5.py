# ch19_5.py
import openpyxl

fn = 'sales.xlsx'
wb = openpyxl.load_workbook(fn)
#ws = wb.get_sheet_by_name('2020Q1')     # 目前工作表2020Q1
ws = wb['2020Q1']
print("儲存格A1 = ", ws['A1'].value)    # A1 
print("儲存格A2 = ", ws['A2'].value)    # A2
print("儲存格B2 = ", ws['B2'].value)    # B2
print("儲存格B3 = ", ws['B3'].value)    # B3
print("儲存格C5 = ", ws['C5'].value)    # C5


