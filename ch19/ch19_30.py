# ch19_30.py
import openpyxl
from openpyxl.styles import Alignment

wb = openpyxl.Workbook()                              # 建立空白的活頁簿
ws = wb.active                                        # 獲得目前工作表
ws['A1'] = 'Ming-Chi Institute of Technology'
ws['A2'] = 'Ming-Chi Institute of Technology'
ws['A3'] = '明志科大'
ws['C3'] = '機械工程系'
ws.merge_cells('A1:E1')                               # 合併A1:E1儲存格
ws.merge_cells('A3:A5')                               # 合併A1:E1儲存格
ws.merge_cells('C3:E4')                               # 合併C3:E4儲存格
ws['A1'].alignment = Alignment(horizontal='center')   # A1儲存格水平置中
ws['A3'].alignment = Alignment(vertical='center')     # A3儲存格垂直置中
ws['C3'].alignment = Alignment(horizontal='center', vertical='center')
wb.save('out19_30.xlsx')                              # 將活頁簿儲存





