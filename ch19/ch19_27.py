# ch19_27.py
import openpyxl

wb = openpyxl.Workbook()            # 建立空白的活頁簿
ws = wb.active                      # 獲得目前工作表
ws['A1'] = 'Peter'                  # 設定名字Peter           
ws['B1'] = 98
ws['A2'] = 'Janet'                  # 設定名字Janet
ws['B2'] = 79
ws['A3'] = 'Nelson'                 # 設定名字Nelson
ws['B3'] = 81
ws['A4'] = '總分'
ws['B4'] = '=SUM(B1:B3)'            # 計算總分
ws['A5'] = '平均'
ws['B5'] = '=AVERAGE(B1:B3)'        # 計算平均
ws['A6'] = '最高分'
ws['B6'] = '=MAX(B1:B3)'            # 計算最高分
ws['A7'] = '最低分'
ws['B7'] = '=MIN(B1:B3)'            # 計算最低分
wb.save('out19_27.xlsx')            # 將活頁簿儲存





