# ch19_15.py
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

fn = 'sales.xlsx'
wb = openpyxl.load_workbook(fn, data_only=True)
#ws = wb.get_sheet_by_name('2020Q1')             # 目前工作表2020Q1
ws = wb['2020Q1']
print("欄數= ",get_column_letter(ws.max_column))
print("3   = ",get_column_letter(3))
print("27  = ",get_column_letter(27))
print("100 = ",get_column_letter(100))
print("800 = ",get_column_letter(800))

print("A   = ", column_index_from_string('A'))
print("E   = ", column_index_from_string('E'))
print("AA  = ", column_index_from_string('AA'))
print("AZ  = ", column_index_from_string('AZ'))
print("AAA = ", column_index_from_string('AAA'))


