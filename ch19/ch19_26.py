# ch19_26.py
import openpyxl
from openpyxl.styles import Font

wb = openpyxl.Workbook()                    # 建立空白的活頁簿
ws = wb.active                              # 獲得目前工作表
fontTitle1 = Font(name='Old English Text MT', size=24, color='0000FF')
ws['A1'].font = fontTitle1
ws['A1'] = 'Ming-Chi Institute of Technology'
fontTitle2 = Font(name='Old English Text MT', size=24, color='FF66FF')
ws['A2'].font = fontTitle2
ws['A2'] = 'Ming-Chi Institute of Technology'
wb.save('out19_26.xlsx')                    # 將活頁簿儲存






