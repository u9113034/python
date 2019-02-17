# ch19_25.py
import openpyxl
from openpyxl.styles import Font

wb = openpyxl.Workbook()                    # 建立空白的活頁簿
ws = wb.active                  # 獲得目前工作表
fontTitle1 = Font(name='微軟正黑體', size=24)
ws['A1'].font = fontTitle1                  
ws['A1'] = '明志科技大學'
fontTitle2 = Font(name='Old English Text MT', size=24, bold=True)
ws['A2'].font = fontTitle2
ws['A2'] = 'Ming-Chi Institute of Technology'
fontTitle3 = Font(size=24, bold=True, italic=True)
ws['A3'].font = fontTitle3
ws['A3'] = 'Ming-Chi Institute of Technology'
wb.save('out19_25.xlsx')                    # 將活頁簿儲存






