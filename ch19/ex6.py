import openpyxl
from openpyxl.styles import Alignment,Font

def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    # Loops through selected Rows
    for i in range(startRow, endRow + 1, 1):
        # Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol, endCol + 1, 1):
            rowSelected.append(sheet.cell(row=i, column=j).value)
        # Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)

    return rangeSelected


# Paste range
# Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving, copiedData):
    countRow = 0
    for i in range(startRow, endRow + 1, 1):
        countCol = 0
        for j in range(startCol, endCol + 1, 1):
            sheetReceiving.cell(row=i, column=j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

wb=openpyxl.load_workbook('ex6.xlsx',data_only=True)
#wb.create_sheet(title='2020Total')
ws=wb['2020Total']
ws.merge_cells('A1:K1')
ws['A1']='銷售業績表'
font_style=Font(name='微軟正黑體',size=20,color='0000FF')
ws['A1'].font=font_style
ws.row_dimensions[1].height=30
ws['A1'].alignment=Alignment(horizontal='center',vertical='center')
ws1=wb['2020Q1']
ws2=wb['2020Q2']
ws3=wb['2020Q3']
selectedRange=copyRange(1,2,4,12,ws1)
pasteRange(1,2,4,12,ws,selectedRange)
selectedRange=copyRange(2,2,4,12,ws2)
pasteRange(5,2,7,12,ws,selectedRange)
selectedRange=copyRange(2,2,4,12,ws3)
pasteRange(8,2,10,12,ws,selectedRange)
ws['K2']='總計'
for j in range(3,13):
    row=ws.cell(row=j,column=11).row
    ws.cell(row=j,column=11).value='=SUM(B'+str(row)+':J'+str(row)+')'
wb.save('ex6.xlsx')