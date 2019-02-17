# ch19_6.py
import openpyxl

fn = 'sales.xlsx'
wb = openpyxl.load_workbook(fn)
#ws = wb.get_sheet_by_name('2020Q1')  # 目前工作表2020Q1
ws = wb['2020Q1']
print("儲存格A1 = ", ws['A1'].column, ws['A1'].row, ws['A1'].coordinate)    
print("儲存格A2 = ", ws['A2'].column, ws['A2'].row, ws['A2'].coordinate)    
print("儲存格B2 = ", ws['B2'].column, ws['B2'].row, ws['B2'].coordinate)    
print("儲存格B3 = ", ws['B3'].column, ws['B3'].row, ws['B3'].coordinate)    
print("儲存格C5 = ", ws['C5'].column, ws['C5'].row, ws['C5'].coordinate) 



