# ch19_31.py
import openpyxl

wb = openpyxl.load_workbook('out19_30.xlsx')    # 開啟活頁簿
ws = wb.active                                  # 獲得目前工作表
ws.unmerge_cells('A1:E1')                       # 取消合併A1:E1儲存格
ws.unmerge_cells('A3:A5')                       # 取消合併A1:E1儲存格
ws.unmerge_cells('C3:E4')                       # 取消合併C3:E4儲存格
wb.save('out19_31.xlsx')                        # 將活頁簿儲存





