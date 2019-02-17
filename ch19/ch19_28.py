# ch19_28.py
import openpyxl
from openpyxl.styles import Alignment

wb = openpyxl.Workbook()                # 建立空白的活頁簿
ws = wb.active                          # 獲得目前工作表
ws['A1'] = '深石數位'
ws['B2'] = 'Deep Stone'
ws.row_dimensions[1].height = 40        # 高度是40pt
ws.column_dimensions['B'].width = 20    # 寬度是20個英文字元寬
wb.save('out19_28.xlsx')                # 將活頁簿儲存





