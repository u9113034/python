# ch19_29.py
import openpyxl
from openpyxl.styles import Alignment

wb = openpyxl.Workbook()                # 建立空白的活頁簿
ws = wb.active              # 獲得目前工作表
ws['A1'] = '台科大'
ws['B1'] = '明志科大'
ws['C1'] = '北科大'
ws.row_dimensions[1].height = 40        # 高度是40pt
ws.column_dimensions['B'].width = 20    # 寬度是20個英文字元寬
ws['A1'].alignment = Alignment(horizontal='left', vertical='top')   
ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
ws['C1'].alignment = Alignment(horizontal='right', vertical='bottom')
wb.save('out19_29.xlsx')                # 將活頁簿儲存





