import openpyxl
from openpyxl.styles import Font

wb=openpyxl.load_workbook('ex4-1.xlsx',data_only=True)
#wb=openpyxl.Workbook()
#wb._data_only=True
ws=wb.active
'''Rows=[
    ['','國文','英文','數學','物理','化學'],
    ['張三',89,90,92,78,75],
    ['李四',89,67,78,58,94],
    ['洪五',77,88,99,90,69],
    ['陳六',98,56,77,88,91],
]
wb.title='mywork'
ws['A1']='National Yunlin University of Science and Technology'
fontTitle1=Font(name='Old English Text MT',size=18)
ws['A1'].font=fontTitle1
for row in Rows:
    ws.append(row)
ws['A7']='最高分'
for i in range(2,7):
    col=ws.cell(row=7,column=i).column
    ws.cell(row=7,column=i).value='=MAX('+col+'3:'+col+'6)'

ws['G2']='總分'
for j in range(3,7):
    row=ws.cell(row=j,column=7).row
    ws.cell(row=j,column=7).value='=SUM(B'+str(row)+':F'+str(row)+')'

ws['H2']='第一名'
ws['G7']='=MAX(G3:G6)'
'''
tmp=ws['I2'].value
r=1
for cell in list(ws.columns)[6]:
    if cell.value==tmp:
        r=cell.row
ws['H3']=ws.cell(row=r,column=1).value

wb.save('ex4.xlsx')