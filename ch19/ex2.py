import openpyxl
from openpyxl.styles import Font

wb=openpyxl.Workbook()
ws=wb.active
Rows=[
    ['','國文','英文','數學','物理','化學'],
    ['張三',89,90,92,78,75],
    ['李四',89,67,78,58,94],
    ['洪五',77,88,99,90,69],
    ['陳六',98,56,77,88,91],
]
ws.title='mywork'
ws['A1']='National Yunlin University of Science and Technology'
fontTitle1=Font(name='Old English Text MT',size=18)
ws['A1'].font=fontTitle1
for row in Rows:
    ws.append(row)

wb.save('ex2.xlsx')