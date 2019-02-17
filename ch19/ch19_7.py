# ch19_7.py
import openpyxl

fn = 'sales.xlsx'
wb = openpyxl.load_workbook(fn)
#ws = wb.get_sheet_by_name('2020Q1')  # 目前工作表2020Q1
ws = wb['2020Q1']
print("工作表欄數 = ", ws.max_column)    
print("工作表行數 = ", ws.max_row)    



