# ch19_22.py
import openpyxl

wb = openpyxl.Workbook()                    # 建立空白的活頁簿
ws = wb.active                                # 獲得目前工作表
ws['A1'] = 'Python'
ws['A2'] = 100
wb.save('out19_22.xlsx')                    # 將活頁簿儲存






