import openpyxl

wb=openpyxl.load_workbook('ex5.xlsx',data_only=True)
ws=wb['2020Q1']
ws['A13']='銷售總計'
sum=0
for c in range(2,6):
    for r in range(3,13):
        sum+=int(ws.cell(row=r,column=c).value)
    ws.cell(row=13,column=c).value=sum
    sum=0

wb.save('ex5.xlsx')